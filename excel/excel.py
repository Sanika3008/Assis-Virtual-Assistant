

from asyncio.windows_events import INFINITE
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter

# wb = load_workbook('D:\ecel\Book1.xlsx')
# self.ws = wb.active
class excel:
    def __init__(self) -> None:
        
        self.s={}
        self.list =[]
        self.list2 = []

    def load(self,wokbook):
        w = load_workbook(wokbook)
        self.ws = w.active
    def get_name_and_row_num(self,name):
        
        for i in range(2,INFINITE):
            char = self.ws['A' + str(i)]
            if char.value == None:
                break
            elif char.value[:len(name)] == name:
                self.s.update({i:char.value})
                
        return self.s
    def title(self):
        
        for i in range(1,INFINITE):
            alphabet = get_column_letter(i)
            char = self.ws[alphabet + '1']
            if char.value == None:
                break
            else:
                self.list.append(char.value)

        return self.list
                
    def get_columns_values(self):
        for i in self.s.keys():
            for j in range(2,INFINITE):
                alphabet = get_column_letter(j)
                charr = self.ws[alphabet + str(i)]
                if charr.value == None:
                    break
                else:
                    self.list2.append(charr.value)
        return self.list2
    def show_columns_with_values(self):
        for content in self.list:
            print(content,end=" |              ")
    
        print()
        l=0
        for i in self.s.keys():
            print(self.s[i],end="|        ")
            for j in range(2,INFINITE):
                alphabet = get_column_letter(j)
                charr = self.ws[alphabet + str(i)]
                
                if charr.value == None:
                    break
                else:
                    print(charr.value , end="|               ")
            print()




# s =excel()
# print(s.get_name_and_row_num("Sanika"))
# print(s.title())
# print(s.get_columns_values())
# s.show_columns_with_values()
